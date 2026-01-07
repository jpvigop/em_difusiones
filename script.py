import re
import sys
from pathlib import Path

import pandas as pd

# -----------------------------
# Config (ajustá solo esto)
# -----------------------------
CANDIDATOS_FILE = "archivos/enviarmatisse.xlsx"
EXCLUIR_FILE = "archivos/excluirmatisse.xlsx"

# Nombre de la campaña (para los archivos de salida)
CAMPAIGN_NAME = "matisse7-1"

# Carpeta de salida (se crea automáticamente si no existe)
OUTPUT_FOLDER = "output"

# Columnas en candidatos
COL_CODIGO = "Código/N°"
COL_NOMBRE = "Nombre"
COL_SALIDAS = "Salidas"
COL_VENTAS = "Ventas"
COL_CELULAR = "Celular"
COL_MAIL = "Mail"

NORMALIZE_UY = True

# Filtros
MIN_SALIDAS = 1
MAX_SALIDAS = 2
MIN_VENTAS = None
MAX_VENTAS = None
MAX_DESCUENTO_PERMITIDO = 9


# -----------------------------
# Helpers
# -----------------------------
def excel_nompropio_first_word(name: str) -> str:
    if name is None:
        return ""
    s = str(name).strip()
    if not s:
        return ""
    first = s.split()[0]
    return first.lower().capitalize()


def digits_only(s: str) -> str:
    return re.sub(r"\D", "", s or "")


def extract_discount_percentage(name: str) -> tuple[float | None, str | None]:
    if name is None:
        return None, None
    s = str(name)
    match = re.search(r'(\d+(?:[,\.]\d+)?)\s*%', s)
    if not match:
        return None, None
    pct_str = match.group(1).replace(',', '.')
    matched_text = match.group(0)
    try:
        return float(pct_str), matched_text
    except ValueError:
        return None, None


def split_into_phones(text: str) -> list[str]:
    """
    Divide un texto en teléfonos individuales usando analisis agudo de digitos.
    Ignora basura, espacios y separadores.
    """
    if pd.isna(text) or text == "":
        return []
    s = str(text).strip()
    if not s:
        return []
    
    # 1. Limpieza preliminar
    # Quitar (.0) de excel
    s = re.sub(r'\.0\b', '', s)
    # Normalizar parentesis (quitar contenido)
    s = re.sub(r'\([^)]*\)', '', s)
    
    # 2. Obtener solo digitos
    digits = re.sub(r"\D", "", s)
    
    if len(digits) < 7:
        return []

    phones = []
    rest = digits
    
    # Parser voraz (Greedy) para detectar numeros de Uruguay
    while len(rest) >= 8:
        # Caso 1: 598xxxxxxxxx (11 digitos)
        if rest.startswith("598") and len(rest) >= 11:
            if rest[3] in "924":
                phones.append(rest[:11])
                rest = rest[11:]
                continue
                
        # Caso 2: 0xxxxxxxx (9 digitos)
        if rest.startswith("0") and len(rest) >= 9:
            if rest[1] in "924":
                phones.append(rest[:9])
                rest = rest[9:]
                continue
                
        # Caso 3: xxxxxxxx (8 digitos)
        if rest[0] in "924" and len(rest) >= 8:
            phones.append(rest[:8])
            rest = rest[8:]
            continue
            
        # Si no encaja en ningun patron conocido de UY, descartamos 1 digito y reintentamos
        rest = rest[1:]
        
    return phones


def normalize_uy(phone: str) -> str | None:
    d = digits_only(phone)
    if len(d) < 7:
        return None
    if not NORMALIZE_UY:
        return d
    if d.startswith("00598"):
        d = d[2:]
    if d.startswith("598"):
        return d
    if d.startswith("09") and len(d) == 9:
        return "598" + d[1:]
    if d.startswith("9") and len(d) == 8:
        return "598" + d
    if d.startswith("2") and len(d) == 8:
        return "598" + d
    if d.startswith("4") and len(d) == 8:
        return "598" + d
    return d


def read_all_sheets_excel(path: str) -> list[pd.DataFrame]:
    xls = pd.ExcelFile(path)
    return [pd.read_excel(path, sheet_name=sn) for sn in xls.sheet_names]


def build_exclusion_sets(excluir_path: str) -> tuple[set[str], set[str], set[str]]:
    dfs = read_all_sheets_excel(excluir_path)
    excl_raw = set()
    excl_digits = set()
    excl_norm = set()

    for df in dfs:
        for col in df.columns:
            for val in df[col].dropna().astype(str).tolist():
                phones = split_into_phones(val)
                for phone in phones:
                    p_clean = phone.strip()
                    if not p_clean: continue
                    excl_raw.add(p_clean)
                    
                    d = digits_only(p_clean)
                    if len(d) >= 7:
                        excl_digits.add(d)
                        
                    n = normalize_uy(p_clean)
                    if n:
                        excl_norm.add(n)
    return excl_raw, excl_digits, excl_norm


def find_column(df: pd.DataFrame, candidates: list[str]) -> str | None:
    df_cols_clean = {c.strip().lower(): c for c in df.columns}
    for cand in candidates:
        cand_clean = cand.strip().lower()
        if cand_clean in df_cols_clean:
            return df_cols_clean[cand_clean]
    return None


def apply_salidas_ventas_filters(df: pd.DataFrame, col_salidas: str | None, col_ventas: str | None) -> tuple[pd.DataFrame, int]:
    result = df.copy()
    excluded_count = 0
    
    if col_salidas and col_salidas in result.columns:
        salidas_num = pd.to_numeric(result[col_salidas], errors='coerce')
        mask = pd.Series(True, index=result.index)
        
        if MIN_SALIDAS is not None:
            mask &= (salidas_num >= MIN_SALIDAS)
        if MAX_SALIDAS is not None:
            mask &= (salidas_num <= MAX_SALIDAS)
            
        excluded_count += (~mask).sum()
        result = result[mask]
        
    if col_ventas and col_ventas in result.columns:
        ventas_num = pd.to_numeric(result[col_ventas], errors='coerce')
        mask = pd.Series(True, index=result.index)
        
        if MIN_VENTAS is not None:
            mask &= (ventas_num >= MIN_VENTAS)
        if MAX_VENTAS is not None:
            mask &= (ventas_num <= MAX_VENTAS)
            
        excluded_count += (~mask).sum()
        result = result[mask]
        
    return result, excluded_count


def apply_discount_filter(df: pd.DataFrame, col_nombre: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    if MAX_DESCUENTO_PERMITIDO is None:
        return df, pd.DataFrame()
    
    result = df.copy()
    result["__pct"] = None
    result["__pct_match"] = None
    
    for idx in result.index:
        name = result.loc[idx, col_nombre]
        pct, match_text = extract_discount_percentage(name)
        result.loc[idx, "__pct"] = pct
        result.loc[idx, "__pct_match"] = match_text
    
    mask_keep = result["__pct"].isna() | (result["__pct"] <= MAX_DESCUENTO_PERMITIDO)
    
    df_keep = result[mask_keep].drop(columns=["__pct", "__pct_match"])
    df_excluded = result[~mask_keep].copy()
    df_excluded["Motivo_Descuento"] = df_excluded["__pct_match"]
    
    return df_keep, df_excluded.drop(columns=["__pct", "__pct_match"])


def check_phone_excluded(phone_raw: str, excl_raw: set, excl_digits: set, excl_norm: set) -> bool:
    if phone_raw in excl_raw: return True
    p_digits = digits_only(phone_raw)
    if p_digits in excl_digits: return True
    p_norm = normalize_uy(phone_raw)
    if p_norm and p_norm in excl_norm: return True
    return False


def filter_whole_person_by_exclusion(df: pd.DataFrame, col_celular: str, excl_raw: set, excl_digits: set, excl_norm: set) -> tuple[pd.DataFrame, int]:
    """
    Excluye al CONTACTO completo si ALGUNO de sus teléfonos está en la lista negra.
    """
    tmp = df.copy()
    tmp["__orig_idx"] = tmp.index
    tmp[col_celular] = tmp[col_celular].fillna("").astype(str)
    
    # Chequeo por partes (split)
    tmp["__phones"] = tmp[col_celular].apply(split_into_phones)
    exploded = tmp.explode("__phones")
    exploded["telefono_raw"] = exploded["__phones"].fillna("").astype(str).str.strip()
    exploded = exploded[exploded["telefono_raw"] != ""]
    
    exploded["is_excluded"] = exploded.apply(
        lambda r: check_phone_excluded(r["telefono_raw"], excl_raw, excl_digits, excl_norm),
        axis=1
    )
    
    idx_excluded_split = set(exploded[exploded["is_excluded"]]["__orig_idx"])
    
    # Log matches for debugging
    matches_found = exploded[exploded["is_excluded"]]
    print(f"  [DEBUG] Matches encontrados: {len(matches_found)}")
    for _, r in matches_found.head(5).iterrows():
        print(f"    - {r[COL_NOMBRE]}: {r['telefono_raw']}")
    
    df_keep = df[~df.index.isin(idx_excluded_split)].copy()
    
    return df_keep, len(idx_excluded_split)


def get_first_valid_phone(cell_text: str) -> str:
    phones = split_into_phones(cell_text)
    if phones:
        n = normalize_uy(phones[0])
        return n if n else phones[0]
    return ""


def get_all_valid_phones(cell_text: str) -> list[str]:
    """Devuelve todos los teléfonos normalizados de una celda."""
    phones = split_into_phones(cell_text)
    result = []
    for p in phones:
        n = normalize_uy(p)
        result.append(n if n else p)
    return result


def format_598_to_09(phone_598: str) -> str:
    """Convierte formato 598XXXXXXXX a 09XXXXXXX (formato tradicional UY)."""
    if not phone_598:
        return ""
    d = digits_only(phone_598)
    if d.startswith("598") and len(d) == 11:
        # 59899123456 -> 099123456
        return "0" + d[3:]
    # Si no es formato 598, devolver como está
    return d


# -----------------------------
# Main
# -----------------------------
def main():
    print("Iniciando script v2.0...")
    cand_path = Path(CANDIDATOS_FILE)
    exc_path = Path(EXCLUIR_FILE)

    if not cand_path.exists() or not exc_path.exists():
        print("Error: No encuentro los archivos.")
        sys.exit(1)
    
    # Archivo de salida (un solo archivo con 2 hojas)
    output_path = Path(OUTPUT_FOLDER)
    out_file = output_path / f"{CAMPAIGN_NAME}.xlsx"
    
    print(f"Campaña: {CAMPAIGN_NAME}")
    print(f"Salida: {out_file}")

    df_c = pd.read_excel(cand_path)
    # Reset index to ensure unique simple index for filtering
    df_c = df_c.reset_index(drop=True)
    initial_count = len(df_c)

    col_nombre = find_column(df_c, [COL_NOMBRE, "nombre", "Nombre"])
    col_celular = find_column(df_c, [COL_CELULAR, "celular", "Celular", "Telefono"])
    col_codigo = find_column(df_c, [COL_CODIGO, "Codigo", "codigo", "N"])
    col_salidas = find_column(df_c, [COL_SALIDAS, "salidas"])
    col_ventas = find_column(df_c, [COL_VENTAS, "ventas"])
    col_mail = find_column(df_c, [COL_MAIL, "mail", "email", "Email"])

    if not col_nombre or not col_celular:
        print("Error: Faltan columnas Nombre o Celular.")
        sys.exit(1)

    print(f"Procesando {initial_count} registros.")
    print(f"Filtros: Salidas=[{MIN_SALIDAS}, {MAX_SALIDAS}], Descuento max={MAX_DESCUENTO_PERMITIDO}%")

    # 1. Filtro Descuento
    df_c, df_excl_desc = apply_discount_filter(df_c, col_nombre)
    print(f"Excluidos por descuento: {len(df_excl_desc)}")

    # 2. Filtro Salidas/Ventas
    df_c, excl_sv = apply_salidas_ventas_filters(df_c, col_salidas, col_ventas)
    print(f"Excluidos por Salidas/Ventas: {excl_sv}")

    # 3. Exclusión por Teléfono
    excl_raw, excl_digits, excl_norm = build_exclusion_sets(str(exc_path))
    print(f"Cargados {len(excl_norm)} teléfonos de exclusión.")
    
    df_keep, count_excl_phones = filter_whole_person_by_exclusion(df_c, col_celular, excl_raw, excl_digits, excl_norm)
    print(f"Contactos eliminados por coincidencia de teléfono: {count_excl_phones}")

    # 4. Formateo y Dedup
    # Extraer todos los teléfonos de cada cliente
    df_keep["__all_phones"] = df_keep[col_celular].apply(get_all_valid_phones)
    df_keep["telefono_norm"] = df_keep["__all_phones"].apply(lambda x: x[0] if x else "")
    
    # Excluir registros sin teléfono válido (no se puede enviar WhatsApp sin teléfono)
    count_before_phone_filter = len(df_keep)
    df_keep = df_keep[df_keep["telefono_norm"].notna() & (df_keep["telefono_norm"] != "")]
    print(f"Excluidos sin teléfono válido: {count_before_phone_filter - len(df_keep)}")
    
    count_before_dedup = len(df_keep)
    df_keep = df_keep.drop_duplicates(subset=["telefono_norm"], keep=False)
    print(f"Duplicados removidos (eliminados AMBOS con mismo número): {count_before_dedup - len(df_keep)}")

    # Determinar máximo de teléfonos por cliente
    max_phones = df_keep["__all_phones"].apply(len).max() if len(df_keep) > 0 else 1
    max_phones = max(1, max_phones)  # Al menos 1 columna
    
    # Crear columnas de teléfono en ambos formatos (598 y 09)
    for i in range(max_phones):
        # Formato 598 (internacional)
        col_598 = f"Tel{i+1}_598" if max_phones > 1 else "Tel_598"
        df_keep[col_598] = df_keep["__all_phones"].apply(
            lambda phones, idx=i: str(int(float(phones[idx]))) if idx < len(phones) and phones[idx] else ""
        )
        # Formato 09 (local/tradicional)
        col_09 = f"Tel{i+1}_09" if max_phones > 1 else "Tel_09"
        df_keep[col_09] = df_keep["__all_phones"].apply(
            lambda phones, idx=i: format_598_to_09(phones[idx]) if idx < len(phones) and phones[idx] else ""
        )

    # Output envios
    df_keep["Nombre limpio"] = df_keep[col_nombre].apply(excel_nompropio_first_word)

    cols_out = []
    if col_codigo and col_codigo in df_keep.columns: cols_out.append(col_codigo)
    cols_out.append(col_nombre)
    cols_out.append("Nombre limpio")
    if col_salidas and col_salidas in df_keep.columns: cols_out.append(col_salidas)
    if col_ventas and col_ventas in df_keep.columns: cols_out.append(col_ventas)
    if col_mail and col_mail in df_keep.columns: cols_out.append(col_mail)
    
    # Agregar columnas de teléfono (primero todos 598, luego todos 09)
    for i in range(max_phones):
        col_598 = f"Tel{i+1}_598" if max_phones > 1 else "Tel_598"
        cols_out.append(col_598)
    for i in range(max_phones):
        col_09 = f"Tel{i+1}_09" if max_phones > 1 else "Tel_09"
        cols_out.append(col_09)

    df_env_out = df_keep[[c for c in cols_out if c in df_keep.columns]].copy()
    
    # Preparar excluidos
    if len(df_excl_desc) > 0:
        df_excl_desc["Nombre limpio"] = df_excl_desc[col_nombre].apply(excel_nompropio_first_word)
        cols_excl = [col_nombre, "Nombre limpio", col_celular, "Motivo_Descuento"]
        if col_codigo: cols_excl.insert(0, col_codigo)
        df_ex = df_excl_desc[[c for c in cols_excl if c in df_excl_desc.columns]].copy()
        if col_celular in df_ex.columns:
            df_ex[col_celular] = df_ex[col_celular].astype(str)
    else:
        df_ex = pd.DataFrame(columns=["Nombre", "Motivo_Descuento"])
    
    # Guardar en un solo archivo con 2 hojas
    from openpyxl import Workbook
    
    wb = Workbook()
    
    # Hoja 1: Envios
    ws_envios = wb.active
    ws_envios.title = "Envios"
    
    for col_idx, col_name in enumerate(df_env_out.columns, 1):
        ws_envios.cell(row=1, column=col_idx, value=col_name)
    
    # Identificar columnas de teléfono (Tel_598, Tel_09, Tel1_598, Tel1_09, etc.)
    telefono_col_indices = [i + 1 for i, col in enumerate(df_env_out.columns) if col.startswith("Tel")]
    
    for row_idx, row in enumerate(df_env_out.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws_envios.cell(row=row_idx, column=col_idx)
            if col_idx in telefono_col_indices and value:
                cell.value = str(value)
                cell.number_format = '@'
            else:
                cell.value = value
    
    # Hoja 2: Excluidos
    ws_excluidos = wb.create_sheet(title="Excluidos")
    
    for col_idx, col_name in enumerate(df_ex.columns, 1):
        ws_excluidos.cell(row=1, column=col_idx, value=col_name)
    
    for row_idx, row in enumerate(df_ex.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            ws_excluidos.cell(row=row_idx, column=col_idx, value=value)
    
    wb.save(out_file)

    print("="*40)
    print(f"FINAL: {len(df_env_out)} envios en hoja 'Envios'")
    print(f"FINAL: {len(df_ex)} excluidos en hoja 'Excluidos'")
    print(f"Archivo: {out_file}")
    print("="*40)

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print("ERROR FATAL:", e)
        import traceback
        traceback.print_exc()
        sys.exit(1)
